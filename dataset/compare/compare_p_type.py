"""
goyo_extracted（NAIST誤用コーパス）と artificial_data2 の
助詞(P)データを比較する統計分析スクリプト
"""

import openpyxl
import statistics
import math

# ============================================================
# 1. データ読み込み
# ============================================================

def load_goyo_p(path):
    """goyo_extracted から助詞(P)データを読み込む"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[3]) if row[3] is not None else ''
        # 'p/' で始まるもののみ（'ph'=成句は除外）
        if t.startswith('p/'):
            goyo = row[6]
            sei  = row[7]
            if goyo is not None and sei is not None:
                records.append({
                    'type_detail': t,
                    'tok_goyo': str(row[4]) if row[4] else '',
                    'tok_sei':  str(row[5]) if row[5] else '',
                    'goyo': goyo,
                    'sei':  sei,
                })
    wb.close()
    return records


def load_artificial_p(path):
    """artificial_data2 から助詞(P)データを読み込む"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == 'P':
            goyo = row[4]
            sei  = row[5]
            if goyo is not None and sei is not None:
                records.append({
                    'tok_goyo': str(row[2]) if row[2] else '',
                    'tok_sei':  str(row[3]) if row[3] else '',
                    'goyo': goyo,
                    'sei':  sei,
                })
    wb.close()
    return records


# ============================================================
# 2. 助詞サブタイプ分類
# ============================================================

def classify_subtype_goyo(type_detail):
    """goyo_extracted の type_detail 文字列からサブタイプを判定"""
    parts = type_detail.split('/')
    # p/om/X → 脱落
    if len(parts) >= 2 and parts[1] == 'om':
        return '脱落'
    # p/X/ad → 余剰
    if len(parts) >= 3 and parts[2] == 'ad':
        return '余剰'
    # p/X/Y (X,Y どちらも助詞名) → 選択誤り
    return '選択誤り'


def classify_subtype_art(tok_goyo, tok_sei):
    """誤用・正用トークンからサブタイプを判定"""
    g = tok_goyo.strip()
    s = tok_sei.strip()
    phi = {'φ', 'φ', '', 'None', 'ø'}
    if g in phi:
        return '脱落'
    if s in phi:
        return '余剰'
    return '選択誤り'


# ============================================================
# 3. 統計計算
# ============================================================

def calc_stats(lengths, label):
    n = len(lengths)
    mean  = statistics.mean(lengths)
    var_s = statistics.variance(lengths)        # 標本分散
    var_p = statistics.pvariance(lengths)       # 母分散
    std_s = statistics.stdev(lengths)
    med   = statistics.median(lengths)
    mn    = min(lengths)
    mx    = max(lengths)
    rng   = mx - mn

    sorted_l = sorted(lengths)
    q1 = sorted_l[int(n * 0.25)]
    q3 = sorted_l[int(n * 0.75)]
    iqr = q3 - q1

    # ヒストグラム（10文字幅のバケット）
    buckets = {}
    for v in lengths:
        b = (v // 10) * 10
        buckets[b] = buckets.get(b, 0) + 1

    return {
        'label': label, 'n': n,
        'mean': mean, 'median': med,
        'var_s': var_s, 'var_p': var_p, 'std_s': std_s,
        'min': mn, 'max': mx, 'range': rng,
        'q1': q1, 'q3': q3, 'iqr': iqr,
        'buckets': buckets,
    }


def print_stats(s):
    print(f"  件数          : {s['n']}")
    print(f"  平均文字数    : {s['mean']:.2f}")
    print(f"  中央値        : {s['median']:.1f}")
    print(f"  標本分散      : {s['var_s']:.2f}")
    print(f"  母分散        : {s['var_p']:.2f}")
    print(f"  標本標準偏差  : {s['std_s']:.2f}")
    print(f"  最小 / 最大   : {s['min']} / {s['max']}")
    print(f"  レンジ        : {s['range']}")
    print(f"  Q1 / Q3 / IQR : {s['q1']} / {s['q3']} / {s['iqr']}")


def print_hist(s):
    print(f"\n  【文字数ヒストグラム（10文字幅）: {s['label']}】")
    for b in sorted(s['buckets']):
        bar = '■' * (s['buckets'][b] * 30 // s['n'])
        print(f"  {b:3d}〜{b+9:3d}文字 : {s['buckets'][b]:4d}件 {bar}")


def print_subtype(counts, total, label):
    print(f"\n  【助詞サブタイプ内訳: {label}】")
    for name in ['選択誤り', '脱落', '余剰']:
        c = counts.get(name, 0)
        pct = c / total * 100 if total > 0 else 0
        print(f"  {name:<8}: {c:4d}件  ({pct:.1f}%)")


def print_diff(g, a, key, fmt='.2f'):
    gv = g[key]
    av = a[key]
    diff = av - gv
    sign = '+' if diff >= 0 else ''
    print(f"  {key:<14}: goyo={gv:{fmt}}  art={av:{fmt}}  差={sign}{diff:{fmt}}")


# ============================================================
# 4. メイン
# ============================================================

def main():
    goyo_path = 'goyo_extracted.xlsx'
    art_path  = 'artificial data2.xlsx'

    print("データ読み込み中...")
    goyo_recs = load_goyo_p(goyo_path)
    art_recs  = load_artificial_p(art_path)

    # ---- 文字数リスト ----
    goyo_goyo_len = [len(r['goyo']) for r in goyo_recs]
    goyo_sei_len  = [len(r['sei'])  for r in goyo_recs]
    art_goyo_len  = [len(r['goyo']) for r in art_recs]
    art_sei_len   = [len(r['sei'])  for r in art_recs]

    # ---- サブタイプカウント ----
    goyo_sub = {}
    for r in goyo_recs:
        st = classify_subtype_goyo(r['type_detail'])
        goyo_sub[st] = goyo_sub.get(st, 0) + 1

    art_sub = {}
    for r in art_recs:
        st = classify_subtype_art(r['tok_goyo'], r['tok_sei'])
        art_sub[st] = art_sub.get(st, 0) + 1

    # ---- 統計 ----
    gs_g = calc_stats(goyo_goyo_len, 'goyo誤文')
    gs_s = calc_stats(goyo_sei_len,  'goyo正文')
    as_g = calc_stats(art_goyo_len,  'art誤文')
    as_s = calc_stats(art_sei_len,   'art正文')

    sep = '=' * 60

    # ======== 出力 ========
    print(f"\n{sep}")
    print(" NAIST誤用コーパス（goyo_extracted）: 助詞(P)型")
    print(sep)
    print(f"\n▼ 誤文")
    print_stats(gs_g)
    print(f"\n▼ 正文")
    print_stats(gs_s)
    print_subtype(goyo_sub, len(goyo_recs), 'goyo_extracted')

    print(f"\n{sep}")
    print(" 人工データ（artificial_data2）: 助詞(P)型")
    print(sep)
    print(f"\n▼ 誤文")
    print_stats(as_g)
    print(f"\n▼ 正文")
    print_stats(as_s)
    print_subtype(art_sub, len(art_recs), 'artificial_data2')

    print(f"\n{sep}")
    print(" 差分比較（人工データ − goyo）")
    print(sep)
    print("\n▼ 誤文")
    print_diff(gs_g, as_g, 'mean')
    print_diff(gs_g, as_g, 'median', '.1f')
    print_diff(gs_g, as_g, 'var_s')
    print_diff(gs_g, as_g, 'std_s')
    print_diff(gs_g, as_g, 'min', 'd')
    print_diff(gs_g, as_g, 'max', 'd')
    print_diff(gs_g, as_g, 'range', 'd')
    print_diff(gs_g, as_g, 'iqr', 'd')
    print("\n▼ 正文")
    print_diff(gs_s, as_s, 'mean')
    print_diff(gs_s, as_s, 'median', '.1f')
    print_diff(gs_s, as_s, 'var_s')
    print_diff(gs_s, as_s, 'std_s')
    print_diff(gs_s, as_s, 'min', 'd')
    print_diff(gs_s, as_s, 'max', 'd')
    print_diff(gs_s, as_s, 'range', 'd')
    print_diff(gs_s, as_s, 'iqr', 'd')

    print(f"\n{sep}")
    print(" サブタイプ割合比較")
    print(sep)
    print(f"\n  {'':8}  {'goyo':>10}  {'art':>10}")
    for name in ['選択誤り', '脱落', '余剰']:
        gc = goyo_sub.get(name, 0)
        ac = art_sub.get(name, 0)
        gp = gc / len(goyo_recs) * 100 if goyo_recs else 0
        ap = ac / len(art_recs)  * 100 if art_recs  else 0
        print(f"  {name:<8}: {gc:4d}件({gp:5.1f}%)  {ac:4d}件({ap:5.1f}%)")

    print(f"\n{sep}")
    print(" 文字数ヒストグラム")
    print(sep)
    print_hist(gs_g)
    print_hist(as_g)
    print_hist(gs_s)
    print_hist(as_s)


if __name__ == '__main__':
    main()
