"""
全型合算 統計量比較スクリプト（全17型対応）
  TSV型 (14型): ADV PRON NEG AUX COL DEM CONJ ADJ STL AD V OM NOM ORD
  XLSX型 (3型): P SEM NOT  ← artificial_data2.xlsx から読み込む
goyo_extracted.xlsx と比較し、全体および型別の統計量を出力する。
合格基準: mean差 ±3以内、var_s差 ±30以内
"""
import sys, csv, statistics, math
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

BASE      = Path(__file__).parent.parent
GOYO_PATH = BASE / 'goyo_extracted.xlsx'
ART2_PATH = BASE / 'artificial data2.xlsx'

# ============================================================
# 型定義
# 各エントリ: (label, source, file_or_None, art_filter_fn, goyo_filter_fn)
#   source      : 'tsv' または 'xlsx'
#   file_or_None: TSVファイル名（tsv/配下）、xlsx型は None
#   art_filter_fn(row) -> bool  ※ row はタプル (No, type, err_tok, cor_tok, err, cor, ...)
#   goyo_filter_fn(t, e, c)    ※ t=type(小文字), e=err_tok, c=cor_tok
# ============================================================

def _is_nokoto(e, c):
    e, c = e.strip(), c.strip()
    if ('の' in e and 'こと' in c) or ('こと' in e and 'の' in c):
        return True
    if e in ('の', 'のが', 'のは', 'のを', 'のに') and not c:
        return True
    return False

def _t(row):
    """行からtype文字列を小文字で取得"""
    return str(row[1]).lower().strip() if row[1] is not None else ''

TYPE_DEFS = [
    # ---- TSV型（14型）----
    ('ADV',  'tsv', 'ADV_v1.tsv',            lambda r: _t(r) == 'adv',
                                              lambda t, *_: t == 'adv'),
    ('PRON', 'tsv', 'PRON_v1_full100.tsv',   lambda r: _t(r) == 'pron',
                                              lambda t, *_: t == 'pron'),
    ('NEG',  'tsv', 'NEG_v1.tsv',            lambda r: _t(r) == 'neg',
                                              lambda t, *_: t.startswith('neg')),
    ('AUX',  'tsv', 'AUX_100_v1.tsv',        lambda r: _t(r) == 'aux',
                                              lambda t, *_: t.startswith('aux')),
    ('COL',  'tsv', 'COL_v1.tsv',            lambda r: _t(r) == 'col',
                                              lambda t, *_: t.startswith('col')),
    ('DEM',  'tsv', 'DEM_100.tsv',           lambda r: _t(r) == 'dem',
                                              lambda t, *_: t.startswith('dem')),
    ('CONJ', 'tsv', 'conj_complete_100.tsv', lambda r: _t(r) == 'conj',
                                              lambda t, *_: t in ('conj', 'conj/part')),
    ('ADJ',  'tsv', 'adj_complete_100.tsv',  lambda r: _t(r) == 'adj',
                                              lambda t, *_: t.startswith('adj')),
    ('STL',  'tsv', 'stl_batch1_50.tsv',     lambda r: _t(r) == 'stl',
                                              lambda t, *_: t.startswith('stl')),
    ('AD',   'tsv', 'ad_batch1_50.tsv',      lambda r: _t(r) == 'ad',
                                              lambda t, *_: t.startswith('ad')),
    ('V',    'tsv', 'v_complete_100.tsv',     lambda r: _t(r) == 'v',
                                              lambda t, *_: t.startswith('v') and not t.startswith('vt')),
    ('OM',   'tsv', 'om_batch1_50.tsv',      lambda r: _t(r) == 'om',
                                              lambda t, *_: t.startswith('om')),
    ('NOM',  'tsv', 'nom_batch1_30 .tsv',    lambda r: _t(r) == 'nom',
                                              lambda t, e, c: t == 'noun/comp' and _is_nokoto(e, c)),
    ('ORD',  'tsv', 'ORD_batch1_25.tsv',     lambda r: _t(r) == 'ord',
                                              lambda t, *_: t.startswith('ord')),
    # ---- XLSX型（3型）----
    ('P',    'xlsx', None, lambda r: str(r[1]) == 'P',
                           lambda t, *_: t.startswith('p/')),
    ('SEM',  'xlsx', None, lambda r: str(r[1]) == 'SEM',
                           lambda t, *_: 'sem' in t),
    ('NOT',  'xlsx', None, lambda r: _t(r).startswith('not'),
                           lambda t, *_: t.startswith('not')),
]

# ============================================================
# データ読み込み
# ============================================================

def load_goyo_all(path):
    """goyo_extracted から全17型分のレコードを読み込む"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    filters = [(label, gf) for label, _, _, _, gf in TYPE_DEFS]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[3]).lower().strip() if row[3] else ''
        e = str(row[4]) if row[4] else ''
        c = str(row[5]) if row[5] else ''
        err_text = row[6]
        cor_text = row[7]
        if err_text is None or cor_text is None:
            continue
        for label, gf in filters:
            if gf(t, e, c):
                records.append({
                    'label':   label,
                    'type':    t,
                    'err_tok': e,
                    'cor_tok': c,
                    'err':     str(err_text),
                    'cor':     str(cor_text),
                })
                break
    wb.close()
    return records


def _make_rec(label, row):
    """xlsx/TSV の行から統一レコードを作成"""
    return {
        'label':   label,
        'type':    str(row[1]) if row[1] is not None else '',
        'err_tok': str(row[2]) if row[2] is not None else '',
        'cor_tok': str(row[3]) if row[3] is not None else '',
        'err':     str(row[4]) if row[4] is not None else '',
        'cor':     str(row[5]) if row[5] is not None else '',
    }


def load_artificial_all():
    """全17型の人工データを読み込む（TSV型→TSVファイル、XLSX型→artificial data2.xlsx）"""
    records = []

    # ---- TSV型 ----
    for label, source, fname, art_filter, _ in TYPE_DEFS:
        if source != 'tsv':
            continue
        tsv_path = BASE / 'tsv' / fname
        try:
            with open(tsv_path, encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter='\t')
                for row in reader:
                    if len(row) < 6:
                        continue
                    if str(row[0]).strip() in ('No', 'no', 'NO'):
                        continue
                    if art_filter(row):
                        records.append(_make_rec(label, row))
        except FileNotFoundError:
            print(f"  WARNING: {fname} が見つかりません")

    # ---- XLSX型 ----
    xlsx_labels = [(label, art_filter) for label, source, _, art_filter, _ in TYPE_DEFS
                   if source == 'xlsx']
    if xlsx_labels:
        try:
            wb = openpyxl.load_workbook(ART2_PATH, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] is None:
                    continue
                for label, af in xlsx_labels:
                    if af(row):
                        records.append(_make_rec(label, row))
                        break
            wb.close()
        except PermissionError:
            print(f"  WARNING: {ART2_PATH.name} が別プロセスで開かれています。P/SEM/NOT型をスキップします。")
        except FileNotFoundError:
            print(f"  WARNING: {ART2_PATH.name} が見つかりません。P/SEM/NOT型をスキップします。")

    return records

# ============================================================
# 統計計算
# ============================================================

def calc_stats(lengths):
    n = len(lengths)
    if n < 2:
        return None
    sl = sorted(lengths)
    mn = statistics.mean(lengths)
    vr = statistics.variance(lengths)
    return {
        'n':      n,
        'mean':   mn,
        'median': statistics.median(lengths),
        'var_s':  vr,
        'std_s':  math.sqrt(vr),
        'min':    min(lengths),
        'max':    max(lengths),
        'range':  max(lengths) - min(lengths),
        'q1':     sl[int(n * 0.25)],
        'q3':     sl[int(n * 0.75)],
        'iqr':    sl[int(n * 0.75)] - sl[int(n * 0.25)],
    }

def buckets(lengths, width=5):
    b = {}
    for v in lengths:
        k = (v // width) * width
        b[k] = b.get(k, 0) + 1
    return b

# ============================================================
# 表示ユーティリティ
# ============================================================

SEP  = '=' * 70
SEP2 = '-' * 70
MEAN_TOL  = 3.0
VAR_S_TOL = 30.0

def judge(diff, key):
    tol = MEAN_TOL if key == 'mean' else VAR_S_TOL
    return '✓' if abs(diff) <= tol else '✗'

def print_stats(s, label):
    print(f"\n  ▼ {label}  (n={s['n']})")
    print(f"    平均文字数    : {s['mean']:.2f}")
    print(f"    中央値        : {s['median']:.1f}")
    print(f"    標本分散      : {s['var_s']:.2f}")
    print(f"    標本標準偏差  : {s['std_s']:.2f}")
    print(f"    最小 / 最大   : {s['min']} / {s['max']}")
    print(f"    レンジ        : {s['range']}")
    print(f"    Q1 / Q3 / IQR : {s['q1']} / {s['q3']} / {s['iqr']}")

def print_diff(gs, as_, key, fmt='.2f'):
    gv = gs[key]; av = as_[key]; diff = av - gv
    sign = '+' if diff >= 0 else ''
    ok = ''
    if key in ('mean', 'var_s'):
        ok = f"  {judge(diff, key)}"
    print(f"    {key:<14}: goyo={gv:{fmt}}  art={av:{fmt}}  差={sign}{diff:{fmt}}{ok}")

def print_hist(lengths, label, width=10, bar_max=30, total=None):
    b = buckets(lengths, width)
    n = total or len(lengths)
    print(f"\n  【{label}】")
    for k in sorted(b):
        cnt = b[k]
        bar = '■' * (cnt * bar_max // n)
        pct = cnt / n * 100
        print(f"    {k:3d}〜{k+width-1:3d}文字 : {cnt:5d}件 ({pct:4.1f}%) {bar}")

# ============================================================
# メイン
# ============================================================

def main():
    n_types = len(TYPE_DEFS)
    tsv_types  = sum(1 for _, s, *_ in TYPE_DEFS if s == 'tsv')
    xlsx_types = sum(1 for _, s, *_ in TYPE_DEFS if s == 'xlsx')

    print(SEP)
    print(f'  全型合算 統計量比較（goyo vs 人工データ）  全{n_types}型')
    print(f'  TSV型: {tsv_types}型 / XLSX型: {xlsx_types}型')
    print(f'  合格基準: mean差 ±{MEAN_TOL}以内、var_s差 ±{VAR_S_TOL}以内')
    print(SEP)

    # ---- データ読み込み ----
    print('\ngoyo_extracted.xlsx を読み込み中...')
    g_recs = load_goyo_all(GOYO_PATH)
    print(f'  goyo 合計: {len(g_recs)}件')

    print('人工データを読み込み中...')
    a_recs = load_artificial_all()
    print(f'  人工 合計: {len(a_recs)}件')

    # ---- 型別件数 ----
    print(f'\n{SEP}')
    print('  型別件数')
    print(SEP2)
    a_by_label = {}
    for r in a_recs:
        a_by_label[r['label']] = a_by_label.get(r['label'], 0) + 1

    print(f"  {'型':<6}  {'goyo':>6}  {'人工':>6}  ソース")
    print(SEP2)
    for label, source, fname, *_ in TYPE_DEFS:
        g_n = sum(1 for r in g_recs if r['label'] == label)
        a_n = a_by_label.get(label, 0)
        src_info = fname if source == 'tsv' else 'artificial data2.xlsx'
        print(f"  {label:<6}  {g_n:>6}  {a_n:>6}  {src_info}")
    print(SEP2)
    print(f"  {'合計':<6}  {len(g_recs):>6}  {len(a_recs):>6}")

    # ---- 全体文字数リスト ----
    g_err_len = [len(r['err']) for r in g_recs]
    g_cor_len = [len(r['cor']) for r in g_recs]
    a_err_len = [len(r['err']) for r in a_recs]
    a_cor_len = [len(r['cor']) for r in a_recs]

    gs_err = calc_stats(g_err_len)
    gs_cor = calc_stats(g_cor_len)
    as_err = calc_stats(a_err_len)
    as_cor = calc_stats(a_cor_len)

    # ---- 詳細統計 ----
    print(f'\n{SEP}')
    print(f'  goyo 詳細統計（全{n_types}型合算）')
    print(SEP)
    print_stats(gs_err, '誤文')
    print_stats(gs_cor, '正文')

    print(f'\n{SEP}')
    print(f'  人工データ 詳細統計（全{n_types}型合算）')
    print(SEP)
    print_stats(as_err, '誤文')
    print_stats(as_cor, '正文')

    # ---- 差分比較 ----
    print(f'\n{SEP}')
    print('  差分比較（人工データ − goyo）')
    print(f'  ※ 合格基準: mean差 ±{MEAN_TOL}以内、var_s差 ±{VAR_S_TOL}以内')
    print(SEP)
    print('\n  ▼ 誤文')
    for key, fmt in [('mean','.2f'),('median','.1f'),('var_s','.2f'),
                     ('std_s','.2f'),('min','d'),('max','d'),('range','d'),('iqr','d')]:
        print_diff(gs_err, as_err, key, fmt)
    print('\n  ▼ 正文')
    for key, fmt in [('mean','.2f'),('median','.1f'),('var_s','.2f'),
                     ('std_s','.2f'),('min','d'),('max','d'),('range','d'),('iqr','d')]:
        print_diff(gs_cor, as_cor, key, fmt)

    # ---- 判定サマリー ----
    print(f'\n{SEP}')
    print('  全体判定サマリー')
    print(SEP)
    checks = [
        ('誤文 mean',  gs_err['mean'],  as_err['mean'],  'mean'),
        ('誤文 var_s', gs_err['var_s'], as_err['var_s'], 'var_s'),
        ('正文 mean',  gs_cor['mean'],  as_cor['mean'],  'mean'),
        ('正文 var_s', gs_cor['var_s'], as_cor['var_s'], 'var_s'),
    ]
    passed = 0
    for name, gv, av, key in checks:
        d = av - gv
        j = judge(d, key)
        if j == '✓':
            passed += 1
        s = f"+{d:.2f}" if d >= 0 else f"{d:.2f}"
        print(f"  {name:<12}: goyo={gv:.2f}  art={av:.2f}  差={s}  {j}")
    print(f"\n  合格: {passed}/4")

    # ---- 文字数ヒストグラム（10文字幅）----
    print(f'\n{SEP}')
    print('  文字数ヒストグラム（10文字幅）')
    print(SEP)
    print_hist(g_err_len, 'goyo 誤文', width=10)
    print_hist(a_err_len, '人工 誤文', width=10)
    print_hist(g_cor_len, 'goyo 正文', width=10)
    print_hist(a_cor_len, '人工 正文', width=10)

    # ---- 型別 誤文 mean 比較 ----
    print(f'\n{SEP}')
    print('  型別 誤文 mean 比較')
    print(SEP)
    print(f"  {'型':<6}  {'goyo_n':>6}  {'art_n':>5}  {'goyo_mean':>9}  {'art_mean':>9}  {'差':>8}  判定")
    print(SEP2)
    for label, source, fname, *_ in TYPE_DEFS:
        g_sub = [len(r['err']) for r in g_recs if r['label'] == label]
        a_sub = [len(r['err']) for r in a_recs if r['label'] == label]
        g_n = len(g_sub)
        a_n = len(a_sub)
        if not g_sub or not a_sub:
            note = '(goyo なし)' if not g_sub else '(人工データなし)'
            print(f"  {label:<6}  {g_n:>6}  {a_n:>5}  {'---':>9}  {'---':>9}  {note}")
            continue
        gm = statistics.mean(g_sub)
        am = statistics.mean(a_sub)
        d  = am - gm
        j  = judge(d, 'mean')
        s  = f"+{d:.2f}" if d >= 0 else f"{d:.2f}"
        print(f"  {label:<6}  {g_n:>6}  {a_n:>5}  {gm:>9.2f}  {am:>9.2f}  {s:>8}  {j}")

    print(f'\n{SEP}')
    print('  評価完了')
    print(SEP)


if __name__ == '__main__':
    main()
