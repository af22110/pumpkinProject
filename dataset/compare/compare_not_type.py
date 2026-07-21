"""
goyo_extracted（NAIST誤用コーパス）と artificial_data2 の
表記誤り(NOT)データを比較する統計分析スクリプト

サブタイプ分類（タイプコードに基づく）:
  not/hg  : ひらがな表記誤り
  not/kj  : 漢字表記誤り
  not/smb : 記号・符号誤り
  その他  : not/om/smb など上記以外
"""

import sys
import openpyxl
import statistics
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# 1. サブタイプ分類
# ============================================================

def classify_not_subtype_goyo(type_str):
    t = str(type_str).lower()
    if t == 'not/hg':   return 'not/hg'
    if t == 'not/kj':   return 'not/kj'
    if 'smb' in t:      return 'not/smb'
    return 'その他'

def classify_not_subtype_art(type_str):
    t = str(type_str).lower()
    if t == 'not/hg':   return 'not/hg'
    if t == 'not/kj':   return 'not/kj'
    if 'smb' in t:      return 'not/smb'
    return 'その他'


# ============================================================
# 2. データ読み込み
# ============================================================

def load_goyo_not(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[3]).lower() if row[3] else ''
        if t.startswith('not'):
            goyo = row[6]
            sei  = row[7]
            if goyo is not None and sei is not None:
                records.append({
                    'type_detail': str(row[3]),
                    'tok_goyo': str(row[4]) if row[4] else '',
                    'tok_sei':  str(row[5]) if row[5] else '',
                    'goyo': goyo,
                    'sei':  sei,
                })
    wb.close()
    return records

def load_artificial_not(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[1]).lower() if row[1] else ''
        if t.startswith('not'):
            goyo = row[4]
            sei  = row[5]
            if goyo is not None and sei is not None:
                records.append({
                    'type_detail': str(row[1]),
                    'tok_goyo': str(row[2]) if row[2] else '',
                    'tok_sei':  str(row[3]) if row[3] else '',
                    'goyo': goyo,
                    'sei':  sei,
                })
    wb.close()
    return records


# ============================================================
# 3. 統計計算
# ============================================================

def calc_stats(lengths):
    n = len(lengths)
    sl = sorted(lengths)
    return {
        'n':      n,
        'mean':   statistics.mean(lengths),
        'median': statistics.median(lengths),
        'var_s':  statistics.variance(lengths),
        'std_s':  statistics.stdev(lengths),
        'min':    min(lengths),
        'max':    max(lengths),
        'range':  max(lengths) - min(lengths),
        'q1':     sl[int(n * 0.25)],
        'q3':     sl[int(n * 0.75)],
        'iqr':    sl[int(n * 0.75)] - sl[int(n * 0.25)],
        'buckets': _buckets(lengths),
    }

def _buckets(lengths):
    b = {}
    for v in lengths:
        k = (v // 10) * 10
        b[k] = b.get(k, 0) + 1
    return b


# ============================================================
# 4. 表示ユーティリティ
# ============================================================

SEP = '=' * 60

def print_stats(s, label):
    print(f"\n▼ {label}")
    print(f"  件数          : {s['n']}")
    print(f"  平均文字数    : {s['mean']:.2f}")
    print(f"  中央値        : {s['median']:.1f}")
    print(f"  標本分散      : {s['var_s']:.2f}")
    print(f"  標本標準偏差  : {s['std_s']:.2f}")
    print(f"  最小 / 最大   : {s['min']} / {s['max']}")
    print(f"  レンジ        : {s['range']}")
    print(f"  Q1 / Q3 / IQR : {s['q1']} / {s['q3']} / {s['iqr']}")

def print_subtype(counts, total, label):
    print(f"\n  【NOTサブタイプ内訳: {label}】")
    for name in ['not/hg', 'not/kj', 'not/smb', 'その他']:
        c   = counts.get(name, 0)
        pct = c / total * 100 if total > 0 else 0
        print(f"  {name:<10}: {c:4d}件  ({pct:.1f}%)")

def print_diff(g, a, key, fmt='.2f'):
    gv, av = g[key], a[key]
    diff = av - gv
    sign = '+' if diff >= 0 else ''
    print(f"  {key:<14}: goyo={gv:{fmt}}  art={av:{fmt}}  差={sign}{diff:{fmt}}")

def print_subtype_compare(g_sub, g_n, a_sub, a_n):
    print(f"\n{SEP}")
    print(" サブタイプ割合比較")
    print(SEP)
    print(f"\n  {'':10}  {'goyo':>14}  {'art':>14}  {'差':>8}")
    for name in ['not/hg', 'not/kj', 'not/smb', 'その他']:
        gc  = g_sub.get(name, 0)
        ac  = a_sub.get(name, 0)
        gp  = gc / g_n * 100 if g_n > 0 else 0
        ap  = ac / a_n * 100 if a_n > 0 else 0
        diff = ap - gp
        sign = '+' if diff >= 0 else ''
        print(f"  {name:<10}: {gc:4d}件({gp:5.1f}%)  {ac:4d}件({ap:5.1f}%)  {sign}{diff:.1f}pt")

def print_hist(s, label):
    print(f"\n  【文字数ヒストグラム（10文字幅）: {label}】")
    for b in sorted(s['buckets']):
        bar = '■' * (s['buckets'][b] * 30 // s['n'])
        print(f"  {b:3d}〜{b+9:3d}文字 : {s['buckets'][b]:4d}件 {bar}")


# ============================================================
# 5. メイン
# ============================================================

def main():
    base      = Path(__file__).parent.parent
    goyo_path = base / 'goyo_extracted.xlsx'
    art_path  = base / 'artificial data2.xlsx'

    print("データ読み込み中...")
    goyo_recs = load_goyo_not(goyo_path)
    art_recs  = load_artificial_not(art_path)

    goyo_goyo_len = [len(r['goyo']) for r in goyo_recs]
    goyo_sei_len  = [len(r['sei'])  for r in goyo_recs]
    art_goyo_len  = [len(r['goyo']) for r in art_recs]
    art_sei_len   = [len(r['sei'])  for r in art_recs]

    goyo_sub = {}
    for r in goyo_recs:
        st = classify_not_subtype_goyo(r['type_detail'])
        goyo_sub[st] = goyo_sub.get(st, 0) + 1

    art_sub = {}
    for r in art_recs:
        st = classify_not_subtype_art(r['type_detail'])
        art_sub[st] = art_sub.get(st, 0) + 1

    gs_g = calc_stats(goyo_goyo_len)
    gs_s = calc_stats(goyo_sei_len)
    as_g = calc_stats(art_goyo_len)
    as_s = calc_stats(art_sei_len)

    print(f"\n{SEP}")
    print(" NAIST誤用コーパス（goyo_extracted）: 表記誤り(NOT)型")
    print(SEP)
    print_stats(gs_g, '誤文')
    print_stats(gs_s, '正文')
    print_subtype(goyo_sub, len(goyo_recs), 'goyo_extracted')

    print(f"\n{SEP}")
    print(" 人工データ（artificial_data2）: 表記誤り(NOT)型")
    print(SEP)
    print_stats(as_g, '誤文')
    print_stats(as_s, '正文')
    print_subtype(art_sub, len(art_recs), 'artificial_data2')

    print(f"\n{SEP}")
    print(" 差分比較（人工データ − goyo）")
    print(SEP)
    print("\n▼ 誤文")
    for key, fmt in [('mean','.2f'),('median','.1f'),('var_s','.2f'),
                     ('std_s','.2f'),('min','d'),('max','d'),('range','d'),('iqr','d')]:
        print_diff(gs_g, as_g, key, fmt)
    print("\n▼ 正文")
    for key, fmt in [('mean','.2f'),('median','.1f'),('var_s','.2f'),
                     ('std_s','.2f'),('min','d'),('max','d'),('range','d'),('iqr','d')]:
        print_diff(gs_s, as_s, key, fmt)

    print_subtype_compare(goyo_sub, len(goyo_recs), art_sub, len(art_recs))

    print(f"\n{SEP}")
    print(" 文字数ヒストグラム")
    print(SEP)
    print_hist(gs_g, 'goyo誤文')
    print_hist(as_g, 'art誤文')
    print_hist(gs_s, 'goyo正文')
    print_hist(as_s, 'art正文')


if __name__ == '__main__':
    main()
