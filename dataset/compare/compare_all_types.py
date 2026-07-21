"""
全型総合評価スクリプト
goyo_extracted.xlsx と 各型TSVファイルの統計比較（mean, var_s）
合格基準: mean差 ±3以内、var_s差 ±30以内
"""
import sys, csv, statistics
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

BASE = Path(__file__).parent.parent
GOYO_PATH = BASE / 'goyo_extracted.xlsx'

# ============================================================
# 型定義：(label, tsv_file, art_type_filter, goyo_filter_fn)
# ============================================================

def _is_nokoto(err_tok, cor_tok):
    e, c = err_tok.strip(), cor_tok.strip()
    if ('の' in e and 'こと' in c) or ('こと' in e and 'の' in c):
        return True
    if e in ('の', 'のが', 'のは', 'のを', 'のに') and not c:
        return True
    return False

TYPE_DEFS = [
    # (表示ラベル, TSVファイル名,          art type文字列,  goyo filter lambda)
    ('ADV',  'ADV_v1.tsv',              'adv',   lambda t, e, c: t == 'adv'),
    ('PRON', 'PRON_v1_full100.tsv',     'pron',  lambda t, e, c: t == 'pron'),
    ('NEG',  'NEG_v1.tsv',              'neg',   lambda t, e, c: t.startswith('neg')),
    ('AUX',  'AUX_100_v1.tsv',          'aux',   lambda t, e, c: t.startswith('aux')),
    ('COL',  'COL_v1.tsv',              'col',   lambda t, e, c: t.startswith('col')),
    ('DEM',  'DEM_100.tsv',             'dem',   lambda t, e, c: t.startswith('dem')),
    ('CONJ', 'conj_complete_100.tsv',   'conj',  lambda t, e, c: t in ('conj', 'conj/part')),
    ('ADJ',  'adj_complete_100.tsv',    'adj',   lambda t, e, c: t.startswith('adj')),
    ('STL',  'stl_batch1_50.tsv',       'stl',   lambda t, e, c: t.startswith('stl')),
    ('AD',   'ad_batch1_50.tsv',        'ad',    lambda t, e, c: t.startswith('ad')),
    ('V',    'v_complete_100.tsv',      'v',     lambda t, e, c: t.startswith('v') and not t.startswith('vt')),
    ('OM',   'om_batch1_50.tsv',        'om',    lambda t, e, c: t.startswith('om')),
    ('NOM',  'nom_batch1_30 .tsv',      'nom',   lambda t, e, c: t == 'noun/comp' and _is_nokoto(e, c)),
    ('ORD',  'ORD_batch1_25.tsv',       'ord',   lambda t, e, c: t.startswith('ord')),
]

# ============================================================
# データ読み込み
# ============================================================

def load_goyo(path, goyo_filter):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[3]).lower().strip() if row[3] else ''
        err_tok = str(row[4]) if row[4] else ''
        cor_tok = str(row[5]) if row[5] else ''
        err_text = row[6]
        cor_text = row[7]
        if err_text is not None and cor_text is not None and goyo_filter(t, err_tok, cor_tok):
            records.append({
                'err': str(err_text), 'cor': str(cor_text),
                'err_tok': err_tok, 'cor_tok': cor_tok, 'type': t,
            })
    wb.close()
    return records


def load_artificial(tsv_path, art_type):
    records = []
    try:
        with open(tsv_path, encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter='\t')
            for row in reader:
                if len(row) < 6:
                    continue
                if row[0] in ('No', 'no', 'NO'):
                    continue
                t = str(row[1]).lower().strip() if row[1] else ''
                if t == art_type:
                    records.append({
                        'no':      row[0],
                        'err_tok': row[2],
                        'cor_tok': row[3],
                        'err':     row[4],
                        'cor':     row[5],
                    })
    except FileNotFoundError:
        pass
    return records

# ============================================================
# 統計計算
# ============================================================

def calc_stats(lengths):
    n = len(lengths)
    if n == 0:
        return None
    return {
        'n':     n,
        'mean':  statistics.mean(lengths),
        'var_s': statistics.variance(lengths) if n >= 2 else 0.0,
    }

# ============================================================
# 合格判定
# ============================================================

MEAN_TOL  = 3.0
VAR_S_TOL = 30.0

def judge(diff, key):
    tol = MEAN_TOL if key == 'mean' else VAR_S_TOL
    return '✓' if abs(diff) <= tol else '✗'

# ============================================================
# 整合性チェック
# ============================================================

def consistency_check(art_recs):
    ok = 0
    ng_list = []
    for r in art_recs:
        et = r['err_tok']
        ct = r['cor_tok']
        es = r['err']
        cs = r['cor']
        cnt = es.count(et)
        expected_cs = es.replace(et, ct, 1)
        issues = []
        if cnt != 1:
            issues.append(f'err_tok出現{cnt}回')
        if cs != expected_cs:
            issues.append('cor_sent不一致')
        if issues:
            ng_list.append((r.get('no', '?'), et, ct, ', '.join(issues)))
        else:
            ok += 1
    return ok, ng_list

# ============================================================
# メイン
# ============================================================

def main():
    SEP = '=' * 80
    SEP2 = '-' * 80

    print(SEP)
    print('  NAIST誤用コーパス 人工データ 全型総合評価')
    print(f'  合格基準: mean差 ±{MEAN_TOL}以内、var_s差 ±{VAR_S_TOL}以内')
    print(SEP)

    # ---- goyo を一括読み込み ----
    print('\ngoyo_extracted.xlsx を読み込み中...')
    wb = openpyxl.load_workbook(GOYO_PATH, read_only=True)
    ws = wb.active
    goyo_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = str(row[3]).lower().strip() if row[3] else ''
        e = str(row[4]) if row[4] else ''
        c = str(row[5]) if row[5] else ''
        err_text = row[6]
        cor_text = row[7]
        if err_text is not None and cor_text is not None:
            goyo_rows.append((t, e, c, str(err_text), str(cor_text)))
    wb.close()
    print(f'goyo: {len(goyo_rows)}行読み込み完了\n')

    # ---- 結果収集 ----
    results = []

    for label, tsv_file, art_type, goyo_filter in TYPE_DEFS:
        tsv_path = BASE / 'tsv' / tsv_file

        # goyo フィルタリング
        g_recs = [(err, cor) for t, e, c, err, cor in goyo_rows if goyo_filter(t, e, c)]

        # 人工データ読み込み
        a_recs = load_artificial(tsv_path, art_type)

        # 整合性チェック
        ok_cnt, ng_list = consistency_check(a_recs)

        # 統計
        gs_err = calc_stats([len(r[0]) for r in g_recs])
        gs_cor = calc_stats([len(r[1]) for r in g_recs])
        as_err = calc_stats([len(r['err']) for r in a_recs])
        as_cor = calc_stats([len(r['cor']) for r in a_recs])

        results.append({
            'label':    label,
            'tsv_file': tsv_file,
            'g_n':      len(g_recs),
            'a_n':      len(a_recs),
            'gs_err':   gs_err,
            'gs_cor':   gs_cor,
            'as_err':   as_err,
            'as_cor':   as_cor,
            'ok_cnt':   ok_cnt,
            'ng_list':  ng_list,
        })

    # ============================================================
    # サマリーテーブル
    # ============================================================
    print(SEP)
    print('  サマリーテーブル（全型一覧）')
    print(SEP)

    hdr = f"{'型':<6} {'goyo':>5} {'art':>5}  {'err_mean':>10}  {'err_var_s':>10}  {'cor_mean':>10}  {'cor_var_s':>10}  {'整合性':>8}  {'判定'}"
    print(hdr)
    print(SEP2)

    total_pass = 0
    total_type = 0

    for r in results:
        label = r['label']
        g_n   = r['g_n']
        a_n   = r['a_n']
        gs_e  = r['gs_err']
        gs_c  = r['gs_cor']
        as_e  = r['as_err']
        as_c  = r['as_cor']
        ok_c  = r['ok_cnt']
        ng    = r['ng_list']

        if not a_n:
            print(f"{label:<6} {g_n:>5} {'---':>5}  {'TSVなし':>52}")
            continue

        # 差分と判定
        flags = []
        if gs_e and as_e:
            d_em = as_e['mean']  - gs_e['mean']
            d_ev = as_e['var_s'] - gs_e['var_s']
            d_cm = as_c['mean']  - gs_c['mean']  if gs_c and as_c else None
            d_cv = as_c['var_s'] - gs_c['var_s'] if gs_c and as_c else None

            j_em = judge(d_em, 'mean')
            j_ev = judge(d_ev, 'var_s')
            j_cm = judge(d_cm, 'mean')  if d_cm is not None else '?'
            j_cv = judge(d_cv, 'var_s') if d_cv is not None else '?'

            flags = [j_em, j_ev, j_cm, j_cv]
            passed = flags.count('✓')
            verdict = '全合格' if passed == 4 else f'{passed}/4合格'

            em_str  = f"{gs_e['mean']:5.1f}→{as_e['mean']:5.1f}({'+' if d_em>=0 else ''}{d_em:.1f}){j_em}"
            ev_str  = f"{gs_e['var_s']:6.1f}→{as_e['var_s']:6.1f}({'+' if d_ev>=0 else ''}{d_ev:.1f}){j_ev}"
            cm_str  = f"{gs_c['mean']:5.1f}→{as_c['mean']:5.1f}({'+' if d_cm>=0 else ''}{d_cm:.1f}){j_cm}" if d_cm is not None else '---'
            cv_str  = f"{gs_c['var_s']:6.1f}→{as_c['var_s']:6.1f}({'+' if d_cv>=0 else ''}{d_cv:.1f}){j_cv}" if d_cv is not None else '---'

            consist_str = f"{ok_c}/{a_n}OK" + (f" NG:{len(ng)}件" if ng else '')

            print(f"{label:<6} {g_n:>5} {a_n:>5}")
            print(f"       誤文 mean: {em_str}")
            print(f"       誤文 var_s: {ev_str}")
            print(f"       正文 mean: {cm_str}")
            print(f"       正文 var_s: {cv_str}")
            print(f"       整合性: {consist_str}  → {verdict}")
            print()

            if passed == 4:
                total_pass += 1
        else:
            print(f"{label:<6} {g_n:>5} {a_n:>5}  (goyo側データなし)")
            print()

        total_type += 1

    # ============================================================
    # 合格/不合格 一覧
    # ============================================================
    print(SEP)
    print('  合格/不合格 一覧（mean差±3、var_s差±30）')
    print(SEP)

    header2 = f"{'型':<6}  {'goyo_n':>6}  {'art_n':>5}  {'err_mean':^14}  {'err_var_s':^14}  {'cor_mean':^14}  {'cor_var_s':^14}  {'整合性':^10}"
    print(header2)
    print(SEP2)

    all_pass_labels = []
    ng_labels = []

    for r in results:
        label = r['label']
        g_n   = r['g_n']
        a_n   = r['a_n']
        gs_e  = r['gs_err']
        gs_c  = r['gs_cor']
        as_e  = r['as_err']
        as_c  = r['as_cor']
        ok_c  = r['ok_cnt']
        ng    = r['ng_list']

        if not a_n or not gs_e:
            print(f"{label:<6}  {g_n:>6}  {a_n if a_n else '---':>5}  {'データ不足':^58}")
            continue

        d_em = as_e['mean']  - gs_e['mean']
        d_ev = as_e['var_s'] - gs_e['var_s']
        d_cm = as_c['mean']  - gs_c['mean']  if gs_c and as_c else None
        d_cv = as_c['var_s'] - gs_c['var_s'] if gs_c and as_c else None

        j_em = judge(d_em, 'mean')
        j_ev = judge(d_ev, 'var_s')
        j_cm = judge(d_cm, 'mean')  if d_cm is not None else '-'
        j_cv = judge(d_cv, 'var_s') if d_cv is not None else '-'

        def fmt_d(d, j):
            if d is None: return f"{'---':^14}"
            s = f"{'+' if d >= 0 else ''}{d:.1f}"
            return f"{s:^12}{j} "

        consist_str = 'OK' if not ng else f'NG:{len(ng)}'
        print(f"{label:<6}  {g_n:>6}  {a_n:>5}  {fmt_d(d_em,j_em)} {fmt_d(d_ev,j_ev)} {fmt_d(d_cm,j_cm)} {fmt_d(d_cv,j_cv)} {consist_str:^10}")

        flags_all = [j_em, j_ev, j_cm, j_cv]
        if all(f == '✓' for f in flags_all if f != '-') and not ng:
            all_pass_labels.append(label)
        else:
            ng_labels.append(label)

    print(SEP2)
    print(f"\n  全合格型 ({len(all_pass_labels)}型): {', '.join(all_pass_labels) if all_pass_labels else 'なし'}")
    print(f"  要確認型 ({len(ng_labels)}型): {', '.join(ng_labels) if ng_labels else 'なし'}")

    # ============================================================
    # 各型の詳細数値
    # ============================================================
    print(f"\n{SEP}")
    print('  各型 詳細数値（goyo vs 人工データ）')
    print(SEP)

    for r in results:
        label = r['label']
        g_n   = r['g_n']
        a_n   = r['a_n']
        gs_e  = r['gs_err']
        gs_c  = r['gs_cor']
        as_e  = r['as_err']
        as_c  = r['as_cor']
        ok_c  = r['ok_cnt']
        ng    = r['ng_list']

        print(f"\n▼ {label}型   goyo:{g_n}件  人工:{a_n}件  ({r['tsv_file']})")
        if not a_n:
            print("  TSVファイルが見つかりません")
            continue
        if not gs_e:
            print("  goyo側にデータなし（人工データのみ）")
            if as_e:
                print(f"  人工誤文  mean={as_e['mean']:.2f}  var_s={as_e['var_s']:.2f}")
                if as_c:
                    print(f"  人工正文  mean={as_c['mean']:.2f}  var_s={as_c['var_s']:.2f}")
            continue

        def row_str(g_val, a_val, key):
            d = a_val - g_val
            j = judge(d, key)
            s = f"+{d:.2f}" if d >= 0 else f"{d:.2f}"
            return f"goyo={g_val:.2f}  art={a_val:.2f}  差={s} {j}"

        print(f"  誤文 mean : {row_str(gs_e['mean'],  as_e['mean'],  'mean')}")
        print(f"  誤文 var_s: {row_str(gs_e['var_s'], as_e['var_s'], 'var_s')}")
        if gs_c and as_c:
            print(f"  正文 mean : {row_str(gs_c['mean'],  as_c['mean'],  'mean')}")
            print(f"  正文 var_s: {row_str(gs_c['var_s'], as_c['var_s'], 'var_s')}")
        consist_str = f"{ok_c}/{a_n}OK"
        if ng:
            consist_str += f"  NG{len(ng)}件"
            for no, et, ct, msg in ng[:5]:
                print(f"    NG No.{no} [{et}→{ct}] {msg}")
        print(f"  整合性    : {consist_str}")

    # ============================================================
    # 定性的評価サマリー
    # ============================================================
    print(f"\n{SEP}")
    print('  定性的評価（各型の特徴と再現度）')
    print(SEP)

    qualitative = {
        'ADV':  '副詞の誤用（程度副詞・時間副詞の混同など）。goyo件数少（12件）のため目標値が不安定だが、文体・文脈の多様性は確保されている。',
        'PRON': '代名詞誤用（複数形・コ/ソ/ア・数量詞）5サブタイプをカバー。正文var_sがやや低いが、err_sent→cor_sent置換制約による構造的制約のため受容可。',
        'NEG':  '否定表現の誤用。「ない/なし/いない」等の混同を含む多様なパターンを生成。',
        'AUX':  '助動詞誤用（テンス・アスペクト・モダリティ）。生成量100件で幅広いサブタイプをカバー。',
        'COL':  '共起・コロケーション誤り。動詞・形容詞・名詞の組み合わせ誤りを含む。',
        'DEM':  '指示詞誤用（コ/ソ/ア系・時間指示）。100件で十分な量と多様性。',
        'CONJ': '接続表現誤用（逆接・順接・因果等の混同）。goyo に conj/part サブタイプも含めて比較。',
        'ADJ':  '形容詞誤用（語形・活用・意味的混同）。60件のgoyoに対し100件生成で十分なカバレッジ。',
        'STL':  '文体誤用（書き言葉/話し言葉混用・敬語レベル誤り）。50件でgoyo305件の代表的なパターンを抽出。',
        'AD':   '副詞的表現・程度表現の誤用。50件。AD型はgoyo最大カテゴリ（1046件）のため50件は部分的カバー。',
        'V':    '動詞の誤用（自他動詞混同・語形選択誤り等）。100件で生成。',
        'OM':   '省略誤り（必要な要素の省略）。50件。',
        'NOM':  'の/こと混同（名詞節マーカー誤り）。30件でgoyo「noun/comp」型の典型的誤りを再現。',
        'ORD':  '語順誤り。25件で基本的な語順誤りパターンを生成。',
    }

    for r in results:
        label = r['label']
        comment = qualitative.get(label, '')
        a_n = r['a_n']
        g_n = r['g_n']
        gs_e = r['gs_err']
        as_e = r['as_err']
        as_c = r['as_cor']

        pass_flags = []
        if gs_e and as_e:
            gs_c = r['gs_cor']
            d_em = as_e['mean']  - gs_e['mean']
            d_ev = as_e['var_s'] - gs_e['var_s']
            pass_flags.append(judge(d_em, 'mean') == '✓')
            pass_flags.append(judge(d_ev, 'var_s') == '✓')
            if gs_c and as_c:
                d_cm = as_c['mean']  - gs_c['mean']
                d_cv = as_c['var_s'] - gs_c['var_s']
                pass_flags.append(judge(d_cm, 'mean') == '✓')
                pass_flags.append(judge(d_cv, 'var_s') == '✓')
        ng_count = len(r['ng_list'])
        stat_ok = f"{sum(pass_flags)}/{len(pass_flags)}統計合格" if pass_flags else '統計比較不可'
        consist_ok = f"整合性{'OK' if ng_count==0 else f'NG:{ng_count}件'}"

        print(f"\n■ {label}型（goyo:{g_n}件, 人工:{a_n}件） [{stat_ok}  {consist_ok}]")
        print(f"  {comment}")

    print(f"\n{SEP}")
    print('  評価完了')
    print(SEP)


if __name__ == '__main__':
    main()
